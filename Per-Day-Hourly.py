import datetime
from datetime import datetime
import os
import re
import time
from tkinter.constants import S
import keyboard
import openpyxl
import pyautogui
import pymsgbox
from keyboard import press
from selenium import webdriver
from selenium.common.exceptions import WebDriverException
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common import by
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from openpyxl import Workbook
import csv

chrome_options = Options()
chrome_options.add_argument("nwapp=C:/Program Files (x86)/LSTV/POS/frontend/KRAKEN_RESTO.exe")

driver = webdriver.Chrome("C:/Users/ASUS/OneDrive/Desktop/Python_Selenium/nwjs-sdk-v0.54.0-win-ia32/chromedriver",options=chrome_options)
driver.maximize_window()

time.sleep(3)
# LOG IN 
driver.find_element_by_xpath("/html/body/ion-app/ion-modal/div/page-setup/ion-content/div[2]/form/ion-card/ion-grid/ion-row[3]/ion-col/ion-item/div[1]/div/ion-input/input").send_keys(Keys.BACKSPACE)
time.sleep(1)
driver.find_element_by_xpath("/html/body/ion-app/ion-modal/div/page-setup/ion-content/div[2]/form/ion-card/ion-grid/ion-row[3]/ion-col/ion-item/div[1]/div/ion-input/input").send_keys(4)
driver.find_element_by_xpath("/html/body/ion-app/ion-modal/div/page-setup/ion-content/div[2]/form/ion-card/ion-grid/ion-row[6]/ion-col[2]/ion-item/div[1]/div/ion-input/input").send_keys(80)
time.sleep(1)
driver.find_element_by_xpath("/html/body/ion-app/ion-modal/div/page-setup/ion-content/div[2]/form/ion-card/ion-grid/ion-row[8]/ion-col/button[2]/span").click()
time.sleep(2)
driver.switch_to.alert.accept() 
time.sleep(5)
driver.find_element_by_xpath("/html/body/ion-app/ion-modal/div/page-security-code/ion-content/div[2]/ion-grid/ion-row[5]/ion-col[2]/input").send_keys("HMP-MOO-NCA-ANC-CCP-ZMO-AOH-NRP-CER-ANK")
time.sleep(1)
driver.find_element_by_xpath("/html/body/ion-app/ion-modal/div/page-security-code/ion-content/div[2]/ion-grid/ion-row[6]/ion-col/button/span").click()
time.sleep(2)
driver.find_element_by_xpath("/html/body/ion-app/ng-component/ion-split-pane/ion-nav/page-login/ion-content/div[2]/ion-grid/ion-row[1]/ion-col/form/ion-card/ion-list/ion-item[1]/div[1]/div/ion-input/input").send_keys("lstv")
driver.find_element_by_xpath("/html/body/ion-app/ng-component/ion-split-pane/ion-nav/page-login/ion-content/div[2]/ion-grid/ion-row[1]/ion-col/form/ion-card/ion-list/ion-item[2]/div[1]/div/ion-input/input").send_keys("lstventures")
time.sleep(1)
driver.find_element_by_xpath("/html/body/ion-app/ng-component/ion-split-pane/ion-nav/page-login/ion-content/div[2]/ion-grid/ion-row[1]/ion-col/form/div/button[1]/span").click()
time.sleep(3)

driver.execute_script("var index = 0;var find_element = false;while (find_element == false){if(document.querySelectorAll('.button-inner')[index].innerText == 'CASH FUND'){find_element = true;    }document.querySelectorAll('.button-inner')[index].innerText;index++;}if(find_element){document.querySelectorAll('.button-inner')[index-1].click();}")
time.sleep(2)
buttons = driver.find_elements_by_class_name('button-inner')
time.sleep(1)
buttons[15].click()
time.sleep(2)
driver.find_element_by_xpath('/html/body/ion-app/ion-modal/div/numpad/ion-content/div[2]/div/ion-grid/ion-row[5]/ion-col[3]/button/span').click()
time.sleep(3)
driver.switch_to.alert.accept()
time.sleep(3)
driver.switch_to.alert.accept()
time.sleep(2)
driver.find_element_by_xpath('/html/body/ion-app/ng-component/ion-split-pane/ion-nav/page-orders/ion-content/div[2]/ion-tabs/ion-tab[1]/page-table-view[2]/ion-content/div[2]/div/button/span').click()
time.sleep(1)
driver.find_element_by_xpath('/html/body/ion-app/ion-alert/div/div[4]/button[2]/span').click()
time.sleep(1)

driver.find_element_by_xpath('/html/body/ion-app/ng-component/ion-split-pane/ion-nav/page-ordering/ion-content/div[2]/div/div[1]/ion-list/ion-item[1]/div[1]/div/ion-label/ion-row/ion-col[1]/button/div[1]/div').click()
driver.find_element_by_xpath('/html/body/ion-app/ng-component/ion-split-pane/ion-nav/page-ordering/ion-content/div[2]/div/div[1]/ion-list/ion-item[1]/div[1]/div/ion-label/ion-row/ion-col[1]/button/div[1]/div').click()
time.sleep(1)
driver.find_element_by_xpath('/html/body/ion-app/ng-component/ion-split-pane/ion-nav/page-ordering/ion-content/div[2]/div/div[1]/ion-list/ion-item[1]/div[1]/div/ion-label/ion-row/ion-col[2]/button/div[1]/div').click()
driver.find_element_by_xpath('/html/body/ion-app/ng-component/ion-split-pane/ion-nav/page-ordering/ion-content/div[2]/div/div[1]/ion-list/ion-item[1]/div[1]/div/ion-label/ion-row/ion-col[2]/button/div[1]/div').click()
time.sleep(1)

driver.find_element_by_xpath('/html/body/ion-app/ng-component/ion-split-pane/ion-nav/page-ordering/ion-footer/ion-toolbar/div[2]/ion-row/ion-col[3]/button/span').click()
time.sleep(2)
driver.switch_to.alert.accept() 
time.sleep(1)
driver.find_element_by_xpath('/html/body/ion-app/ng-component/ion-split-pane/ion-nav/page-ordering/ion-footer/ion-toolbar/div[2]/ion-row/ion-col[6]/button/span').click()    
time.sleep(2)

pos_grossSales = driver.find_element_by_xpath('/html/body/ion-app/ng-component/ion-split-pane/ion-nav/page-order-preview/ion-content/div[2]/div/div[2]/div/div[13]/h4[2]').get_attribute('innerText')
pos_serviceCharge = driver.find_element_by_xpath('/html/body/ion-app/ng-component/ion-split-pane/ion-nav/page-order-preview/ion-content/div[2]/div/div[2]/div/div[12]/h4[2]').get_attribute('innerText')
pos_netSales = float(pos_grossSales) - float(pos_serviceCharge)
pos_vatSales = driver.find_element_by_xpath('/html/body/ion-app/ng-component/ion-split-pane/ion-nav/page-order-preview/ion-content/div[2]/div/div[2]/div/div[14]/h4[2]').get_attribute('innerText')
pos_vatAmount = driver.find_element_by_xpath('/html/body/ion-app/ng-component/ion-split-pane/ion-nav/page-order-preview/ion-content/div[2]/div/div[2]/div/div[15]/h4[2]').get_attribute('innerText')
pos_totalOfTransaction = 1

driver.find_element_by_xpath('/html/body/ion-app/ng-component/ion-split-pane/ion-nav/page-order-preview/ion-content/div[2]/div/div[1]/div/ion-list[4]/div[1]/button/div[1]/div').click()
time.sleep(1)
driver.find_element_by_xpath('/html/body/ion-app/ng-component/ion-split-pane/ion-nav/numpad/ion-content/div[2]/div/ion-grid/ion-row[5]/ion-col[3]/button/span').click()
time.sleep(1)
driver.find_element_by_xpath('/html/body/ion-app/ion-modal/div/page-custom-alert/ion-footer/ion-row/a[2]').click()
time.sleep(3)
driver.switch_to.alert.accept()  
time.sleep(2)
driver.find_element_by_xpath('/html/body/ion-app/ion-alert/div/div[3]/button/span').click()
time.sleep(1)
driver.find_element_by_xpath('/html/body/ion-app/ng-component/ion-split-pane/ion-nav/page-orders/ion-header/ion-navbar/ion-buttons/button').click()
time.sleep(1)
driver.find_element_by_xpath('/html/body/ion-app/ion-popover/div/div[2]/div/popover/ion-list/button[2]/div[1]/div').click()
time.sleep(1)
#CASH DECLARE
driver.find_element_by_xpath('/html/body/ion-app/ng-component/ion-split-pane/ion-nav/page-home/ion-content/div[2]/ion-list/button[4]/span').click()
time.sleep(1)
driver.find_element_by_xpath('/html/body/ion-app/ion-modal/div/page-cash-declaration/ion-content/div[2]/form/ion-grid/ion-row[4]/ion-col[3]/ion-item/div[1]/div/ion-input/input').send_keys(5)
time.sleep(1)
driver.find_element_by_xpath('/html/body/ion-app/ion-modal/div/page-cash-declaration/ion-content/div[2]/form/ion-grid/ion-row[6]/ion-col[1]/button/span').click()
time.sleep(1)
driver.find_element_by_xpath('/html/body/ion-app/ng-component/ion-split-pane/ion-nav/page-home/ion-content/div[2]/ion-list/button[7]/span').click()
time.sleep(1)
driver.find_element_by_xpath('/html/body/ion-app/ng-component/ion-split-pane/ion-nav/page-reports/ion-content/div[2]/ion-list/button[4]/div[1]/div').click()
time.sleep(1)
driver.find_element_by_xpath('/html/body/ion-app/ion-modal/div/page-report-option/ion-content/div[2]/form/ion-grid/ion-card[2]/ion-row[1]/ion-col/ion-item/div[1]/div/ion-select/button/span').click()
time.sleep(1)
driver.find_element_by_xpath('/html/body/ion-app/ion-alert/div/div[3]/div/button[6]/span/div[1]').click()
time.sleep(1)
driver.find_element_by_xpath('/html/body/ion-app/ion-alert/div/div[4]/button[2]/span').click()
time.sleep(1)
driver.find_element_by_xpath('/html/body/ion-app/ion-modal/div/page-report-option/ion-content/div[2]/form/ion-grid/ion-card[2]/ion-row[2]/ion-col[1]/ion-item/div[1]/div/ion-datetime/button/span').click()
time.sleep(1)
driver.find_element_by_xpath('/html/body/ion-app/ion-picker-cmp/div/div[1]/div[2]/button/span').click()
time.sleep(1)
driver.find_element_by_xpath('/html/body/ion-app/ion-modal/div/page-report-option/ion-content/div[2]/form/ion-grid/ion-card[2]/ion-row[2]/ion-col[2]/ion-item/div[1]/div/ion-datetime/button/span').click()
time.sleep(1)
driver.find_element_by_xpath('/html/body/ion-app/ion-picker-cmp/div/div[1]/div[2]/button/span').click()
time.sleep(1)
driver.find_element_by_xpath('/html/body/ion-app/ion-modal/div/page-report-option/ion-content/div[2]/form/ion-grid/ion-card[3]/ion-row[2]/ion-col[2]/ion-item/ion-checkbox/button').click()
time.sleep(1)
driver.find_element_by_xpath('/html/body/ion-app/ion-modal/div/page-report-option/ion-header/ion-navbar/div[2]/ion-title/div/button/span').click()
time.sleep(3)
driver.switch_to.window(driver.window_handles[1])
time.sleep(1)
driver.minimize_window()
time.sleep(1.5)
keyboard.write('p')
time.sleep(1)
press('enter')
time.sleep(2)
keyboard.write('q')
time.sleep(1)
press('enter')
time.sleep(1)
driver.maximize_window()
time.sleep(2)

wb = Workbook()
ws = wb.active
with open('C:/Users/ASUS/Downloads/p.csv', 'r') as f:
    for row in csv.reader(f):
        ws.append(row)
wb.save('pdh.xlsx')
time. sleep(1)
path = "pdh.xlsx"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active        


now = datetime.now()
current_time = now.strftime("%H")


no_of_transactions =  sheet_obj.cell(row = (int(current_time) + 12), column = 3)
totalSales = sheet_obj.cell(row = (int(current_time)+ 12), column = 4)
totalItemsSalesAmount = sheet_obj.cell(row = (int(current_time) + 12), column = 8)

pymsgbox.alert("No. of Transactions: " + str(no_of_transactions.value) + '\n' + "POS No. of Transactions: " + str(pos_totalOfTransaction),timeout=4000) 
if float(no_of_transactions.value) == float(pos_totalOfTransaction):
    pymsgbox.alert('Total No. of Transaction Match ✔️',timeout=3000)
else:
    pymsgbox.alert('Total No. of Transaction Not Match ❌',timeout=3000)    

pymsgbox.alert("Total Sales: " + str(totalSales.value) + '\n' + "POS Total Sales: " + str(pos_grossSales),timeout=4000) 
if float(totalSales.value) == float(pos_grossSales):
    pymsgbox.alert('Total Sales Match ✔️',timeout=3000)
else:
    pymsgbox.alert('Total Sales Not Match ❌',timeout=3000)  

pymsgbox.alert("Total Items Sales Amount: " + str(totalItemsSalesAmount.value) + '\n' + "POS Total Items Sales Amount: " + str(pos_netSales),timeout=4000) 
if float(totalItemsSalesAmount.value) == float(pos_netSales):
    pymsgbox.alert('Total Sales Match ✔️',timeout=3000)
else:
    pymsgbox.alert('Total Sales Not Match ❌',timeout=3000)  
