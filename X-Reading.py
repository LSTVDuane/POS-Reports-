import datetime
import os
import re
import time
from tkinter.constants import S
import keyboard
import openpyxl
import pyautogui
import pymsgbox
from keyboard import press
from selenium import webdriver
from selenium.common.exceptions import WebDriverException
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common import by
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from openpyxl import Workbook
import csv

chrome_options = Options()
chrome_options.add_argument("nwapp=C:/Program Files (x86)/LSTV/POS/frontend/KRAKEN_RESTO.exe")

driver = webdriver.Chrome("C:/Users/ASUS/OneDrive/Desktop/Python_Selenium/nwjs-sdk-v0.54.0-win-ia32/chromedriver",options=chrome_options)
driver.maximize_window()

time.sleep(3)
# LOG IN 
driver.find_element_by_xpath("/html/body/ion-app/ion-modal/div/page-setup/ion-content/div[2]/form/ion-card/ion-grid/ion-row[3]/ion-col/ion-item/div[1]/div/ion-input/input").send_keys(Keys.BACKSPACE)
time.sleep(1)
driver.find_element_by_xpath("/html/body/ion-app/ion-modal/div/page-setup/ion-content/div[2]/form/ion-card/ion-grid/ion-row[3]/ion-col/ion-item/div[1]/div/ion-input/input").send_keys(4)
driver.find_element_by_xpath("/html/body/ion-app/ion-modal/div/page-setup/ion-content/div[2]/form/ion-card/ion-grid/ion-row[6]/ion-col[2]/ion-item/div[1]/div/ion-input/input").send_keys(80)
time.sleep(1)
driver.find_element_by_xpath("/html/body/ion-app/ion-modal/div/page-setup/ion-content/div[2]/form/ion-card/ion-grid/ion-row[8]/ion-col/button[2]/span").click()
time.sleep(2)
driver.switch_to.alert.accept() 
time.sleep(5)
driver.find_element_by_xpath("/html/body/ion-app/ion-modal/div/page-security-code/ion-content/div[2]/ion-grid/ion-row[5]/ion-col[2]/input").send_keys("HMP-MOO-NCA-ANC-CCP-ZMO-AOH-NRP-CER-ANK")
time.sleep(1)
driver.find_element_by_xpath("/html/body/ion-app/ion-modal/div/page-security-code/ion-content/div[2]/ion-grid/ion-row[6]/ion-col/button/span").click()
time.sleep(2)
driver.find_element_by_xpath("/html/body/ion-app/ng-component/ion-split-pane/ion-nav/page-login/ion-content/div[2]/ion-grid/ion-row[1]/ion-col/form/ion-card/ion-list/ion-item[1]/div[1]/div/ion-input/input").send_keys("lstv")
driver.find_element_by_xpath("/html/body/ion-app/ng-component/ion-split-pane/ion-nav/page-login/ion-content/div[2]/ion-grid/ion-row[1]/ion-col/form/ion-card/ion-list/ion-item[2]/div[1]/div/ion-input/input").send_keys("lstventures")
time.sleep(1)
driver.find_element_by_xpath("/html/body/ion-app/ng-component/ion-split-pane/ion-nav/page-login/ion-content/div[2]/ion-grid/ion-row[1]/ion-col/form/div/button[1]/span").click()
time.sleep(3)

driver.execute_script("var index = 0;var find_element = false;while (find_element == false){if(document.querySelectorAll('.button-inner')[index].innerText == 'CASH FUND'){find_element = true;    }document.querySelectorAll('.button-inner')[index].innerText;index++;}if(find_element){document.querySelectorAll('.button-inner')[index-1].click();}")
time.sleep(2)
buttons = driver.find_elements_by_class_name('button-inner')
time.sleep(1)
buttons[15].click()
time.sleep(2)
driver.find_element_by_xpath('/html/body/ion-app/ion-modal/div/numpad/ion-content/div[2]/div/ion-grid/ion-row[5]/ion-col[3]/button/span').click()
time.sleep(3)
driver.switch_to.alert.accept()
time.sleep(3)
driver.switch_to.alert.accept()
time.sleep(2)
driver.find_element_by_xpath('/html/body/ion-app/ng-component/ion-split-pane/ion-nav/page-orders/ion-content/div[2]/ion-tabs/ion-tab[1]/page-table-view[2]/ion-content/div[2]/div/button/span').click()
time.sleep(1)
driver.find_element_by_xpath('/html/body/ion-app/ion-alert/div/div[4]/button[2]/span').click()
time.sleep(1)

driver.find_element_by_xpath('/html/body/ion-app/ng-component/ion-split-pane/ion-nav/page-ordering/ion-content/div[2]/div/div[1]/ion-list/ion-item[1]/div[1]/div/ion-label/ion-row/ion-col[1]/button/div[1]/div').click()
driver.find_element_by_xpath('/html/body/ion-app/ng-component/ion-split-pane/ion-nav/page-ordering/ion-content/div[2]/div/div[1]/ion-list/ion-item[1]/div[1]/div/ion-label/ion-row/ion-col[1]/button/div[1]/div').click()
time.sleep(1)
driver.find_element_by_xpath('/html/body/ion-app/ng-component/ion-split-pane/ion-nav/page-ordering/ion-content/div[2]/div/div[1]/ion-list/ion-item[1]/div[1]/div/ion-label/ion-row/ion-col[2]/button/div[1]/div').click()
driver.find_element_by_xpath('/html/body/ion-app/ng-component/ion-split-pane/ion-nav/page-ordering/ion-content/div[2]/div/div[1]/ion-list/ion-item[1]/div[1]/div/ion-label/ion-row/ion-col[2]/button/div[1]/div').click()
time.sleep(1)

driver.find_element_by_xpath('/html/body/ion-app/ng-component/ion-split-pane/ion-nav/page-ordering/ion-footer/ion-toolbar/div[2]/ion-row/ion-col[3]/button/span').click()
time.sleep(2)
driver.switch_to.alert.accept() 
time.sleep(1)
driver.find_element_by_xpath('/html/body/ion-app/ng-component/ion-split-pane/ion-nav/page-ordering/ion-footer/ion-toolbar/div[2]/ion-row/ion-col[6]/button/span').click()    
time.sleep(2)

pos_grossSales = driver.find_element_by_xpath('/html/body/ion-app/ng-component/ion-split-pane/ion-nav/page-order-preview/ion-content/div[2]/div/div[2]/div/div[13]/h4[2]').get_attribute('innerText')
pos_serviceCharge = driver.find_element_by_xpath('/html/body/ion-app/ng-component/ion-split-pane/ion-nav/page-order-preview/ion-content/div[2]/div/div[2]/div/div[12]/h4[2]').get_attribute('innerText')
pos_netSales = float(pos_grossSales) - float(pos_serviceCharge)
pos_vatSales = driver.find_element_by_xpath('/html/body/ion-app/ng-component/ion-split-pane/ion-nav/page-order-preview/ion-content/div[2]/div/div[2]/div/div[14]/h4[2]').get_attribute('innerText')
pos_vatAmount = driver.find_element_by_xpath('/html/body/ion-app/ng-component/ion-split-pane/ion-nav/page-order-preview/ion-content/div[2]/div/div[2]/div/div[15]/h4[2]').get_attribute('innerText')
pos_vatExempt = driver.find_element_by_xpath('/html/body/ion-app/ng-component/ion-split-pane/ion-nav/page-order-preview/ion-content/div[2]/div/div[2]/div/div[16]/h4[2]').get_attribute('innerText')
pos_totalOfTransaction = 1
pos_pax = 2
pos_quantity = 4
pos_cashFund = 5
pos_cashInDrawer = float(pos_grossSales) + float(pos_cashFund)
pos_posCash = float(pos_cashInDrawer)
pos_cashDeclaration = 5000
pos_shortOver = float(pos_cashDeclaration) - float(pos_posCash) 



driver.find_element_by_xpath('/html/body/ion-app/ng-component/ion-split-pane/ion-nav/page-order-preview/ion-content/div[2]/div/div[1]/div/ion-list[4]/div[1]/button/div[1]/div').click()
time.sleep(1)
driver.find_element_by_xpath('/html/body/ion-app/ng-component/ion-split-pane/ion-nav/numpad/ion-content/div[2]/div/ion-grid/ion-row[5]/ion-col[3]/button/span').click()
time.sleep(1)
driver.find_element_by_xpath('/html/body/ion-app/ion-modal/div/page-custom-alert/ion-footer/ion-row/a[2]').click()
time.sleep(3)
driver.switch_to.alert.accept()  
time.sleep(2)
driver.find_element_by_xpath('/html/body/ion-app/ion-alert/div/div[3]/button/span').click()
time.sleep(1)
driver.find_element_by_xpath('/html/body/ion-app/ng-component/ion-split-pane/ion-nav/page-orders/ion-header/ion-navbar/ion-buttons/button').click()
time.sleep(1)
driver.find_element_by_xpath('/html/body/ion-app/ion-popover/div/div[2]/div/popover/ion-list/button[2]/div[1]/div').click()
time.sleep(1)
#CASH DECLARE
driver.find_element_by_xpath('/html/body/ion-app/ng-component/ion-split-pane/ion-nav/page-home/ion-content/div[2]/ion-list/button[4]/span').click()
time.sleep(1)
driver.find_element_by_xpath('/html/body/ion-app/ion-modal/div/page-cash-declaration/ion-content/div[2]/form/ion-grid/ion-row[4]/ion-col[3]/ion-item/div[1]/div/ion-input/input').send_keys(5)
time.sleep(1)
driver.find_element_by_xpath('/html/body/ion-app/ion-modal/div/page-cash-declaration/ion-content/div[2]/form/ion-grid/ion-row[6]/ion-col[1]/button/span').click()
time.sleep(1)
driver.find_element_by_xpath('/html/body/ion-app/ng-component/ion-split-pane/ion-nav/page-home/ion-content/div[2]/ion-list/button[7]/span').click()
time.sleep(1)
driver.find_element_by_xpath('/html/body/ion-app/ng-component/ion-split-pane/ion-nav/page-reports/ion-content/div[2]/ion-list/button[1]/div[1]/div').click()
time.sleep(1)
driver.find_element_by_xpath('/html/body/ion-app/ion-modal/div/page-xread-filter/ion-content/div[2]/form/ion-card[2]/ion-card-content/ion-row/ion-col/ion-item/div[1]/div/ion-select/button/span').click()
time.sleep(1)
driver.find_element_by_xpath('/html/body/ion-app/ion-alert/div/div[3]/div/button/span/div[1]').click()
time.sleep(1)
driver.find_element_by_xpath('/html/body/ion-app/ion-alert/div/div[4]/button[2]/span').click()
time.sleep(1)
driver.find_element_by_xpath('/html/body/ion-app/ion-modal/div/page-xread-filter/ion-content/div[2]/form/ion-card[3]/ion-row[2]/ion-col[2]/ion-item/ion-checkbox/button').click()
time.sleep(1) 
driver.find_element_by_xpath('/html/body/ion-app/ion-modal/div/page-xread-filter/ion-content/div[2]/form/ion-card[4]/ion-card-content/ion-row/ion-col[1]/button/span').click()
time.sleep(3)
driver.switch_to.window(driver.window_handles[1])
time.sleep(1)
driver.minimize_window()
time.sleep(1)
keyboard.write('p')
time.sleep(1)
press('enter')
time.sleep(1)
driver.maximize_window()
time.sleep(3)

wb = Workbook()
ws = wb.active
with open('C:/Users/ASUS/Downloads/p.csv', 'r') as f:
    for row in csv.reader(f):
        ws.append(row)
wb.save('X-Reading.xlsx')
time. sleep(1)
path = "X-Reading.xlsx"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active        

gross_Sales = sheet_obj.cell(row = 11, column = 3)
serviceCharge = sheet_obj.cell(row = 14, column = 3)    
netSales = sheet_obj.cell(row = 16, column = 3)
vatSales = sheet_obj.cell(row = 17, column = 3)
vatAmount = sheet_obj.cell(row = 18, column = 3)
vatExempt = sheet_obj.cell(row = 19, column = 3)
totalOfTransaction = sheet_obj.cell(row = 20, column = 3)
pax = sheet_obj.cell(row = 21, column = 3)
quantity = sheet_obj.cell(row = 22, column = 3)
cashFund = sheet_obj.cell(row = 28, column = 3)
cashInDrawer = sheet_obj.cell(row = 31, column = 3)
posCash = sheet_obj.cell(row = 32, column = 3)
cashDeclaration = sheet_obj.cell(row = 33, column = 3)
shortOver = sheet_obj.cell(row = 34, column = 3)

pymsgbox.alert('POS Gross Sales: ' + str(pos_grossSales) + '\n' + 'Gross Sales: ' + str(gross_Sales.value),timeout= 4000)
if float(pos_grossSales) == float(gross_Sales.value):
    pymsgbox.alert('Gross Sales Match ✔️',timeout=3000)
else: 
    pymsgbox.alert('Gross Sales Not Match ❌',timeout=3000)

pymsgbox.alert('POS Service Charge: ' + str(pos_serviceCharge) + '\n' + 'Service Charge: ' + str(serviceCharge.value),timeout= 4000)
if float(pos_serviceCharge) == float(serviceCharge.value):
    pymsgbox.alert('Service Charge Match ✔️',timeout=3000)
else: 
    pymsgbox.alert('Service Charge Not Match ❌',timeout=3000)

pymsgbox.alert('POS Net Sales: ' + str(pos_netSales) + '\n' + 'Net Sales: ' + str(netSales.value),timeout= 4000)
if float(pos_netSales) == float(netSales.value):
    pymsgbox.alert('Net Sales Match ✔️',timeout=3000)
else: 
    pymsgbox.alert('Net Sales Not Match ❌',timeout=3000)

pymsgbox.alert('POS Vat Sales: ' + str(pos_vatSales) + '\n' + 'Vat Sales: ' + str(vatSales.value),timeout= 4000)
if float(pos_vatSales) == float(vatSales.value):
    pymsgbox.alert('Vat Sales Match ✔️',timeout=3000)
else: 
    pymsgbox.alert('Vat Sales Not Match ❌',timeout=3000)

pymsgbox.alert('POS Vat Amount: ' + str(pos_vatAmount) + '\n' + 'Vat Amount: ' + str(vatAmount.value),timeout= 4000)
if float(pos_vatAmount) == float(vatAmount.value):
    pymsgbox.alert('Vat Amount Match ✔️',timeout=3000)
else: 
    pymsgbox.alert('Vat Amount Not Match ❌',timeout=3000)

pymsgbox.alert('POS Vat Exempt: ' + str(pos_vatExempt) + '\n' + 'Vat Exempt: ' + str(vatExempt.value),timeout= 4000)
if float(pos_vatExempt) == float(vatExempt.value):
    pymsgbox.alert('Vat Exempt Match ✔️',timeout=3000)
else: 
    pymsgbox.alert('Vat Exempt Not Match ❌',timeout=3000)

pymsgbox.alert('POS No. of Transactions: ' + str(pos_totalOfTransaction) + '\n' + 'No. of Transactions: ' + str(totalOfTransaction.value),timeout= 4000)
if float(pos_totalOfTransaction) == float(totalOfTransaction.value):
    pymsgbox.alert('No. of Transactions Match ✔️',timeout=3000)
else: 
    pymsgbox.alert('No. of Transactionst Not Match ❌',timeout=3000)   

pymsgbox.alert('POS Pax: ' + str(pos_pax) + '\n' + 'Pax: ' + str(pax.value),timeout= 4000)
if float(pos_pax) == float(pax.value):
    pymsgbox.alert('Pax Match ✔️',timeout=3000)
else: 
    pymsgbox.alert('Pax Not Match ❌',timeout=3000)

pymsgbox.alert('POS Quantity: ' + str(pos_quantity) + '\n' + 'Quantity: ' + str(quantity.value),timeout= 4000)
if float(pos_quantity) == float(quantity.value):
    pymsgbox.alert('Quantity Match ✔️',timeout=3000)
else: 
    pymsgbox.alert('Quantity Not Match ❌',timeout=3000)

pymsgbox.alert('POS Cash Fund: ' + str(pos_cashFund) + '\n' + 'Cash Fund: ' + str(cashFund.value),timeout= 4000)
if float(pos_cashFund) == float(cashFund.value):
    pymsgbox.alert('Cash Fund Match ✔️',timeout=3000)
else: 
    pymsgbox.alert('Cash Fund Not Match ❌',timeout=3000)

pymsgbox.alert('POS Cash In Drawer: ' + str(pos_cashInDrawer) + '\n' + 'Cash In Drawer: ' + str(cashInDrawer.value),timeout= 4000)
if float(pos_cashInDrawer) == float(cashInDrawer.value):
    pymsgbox.alert('Cash In Drawer Match ✔️',timeout=3000)
else: 
    pymsgbox.alert('Cash In Drawer Not Match ❌',timeout=3000)    

pymsgbox.alert('POS Cash: ' + str(pos_posCash) + '\n' + 'Cash: ' + str(posCash.value),timeout= 4000)
if float(pos_posCash) == float(posCash.value):
    pymsgbox.alert('Pos Cash Match ✔️',timeout=3000)
else: 
    pymsgbox.alert('Pos Cash Not Match ❌',timeout=3000)   

pymsgbox.alert('POS Cash Declaration: ' + str(pos_cashDeclaration) + '\n' + 'Cash Declaration: ' + str(cashDeclaration.value),timeout= 4000)
if float(pos_cashDeclaration) == float(cashDeclaration.value):
    pymsgbox.alert('Cash Declaration Match ✔️',timeout=3000)
else: 
    pymsgbox.alert('Cash Declaration Not Match ❌',timeout=3000) 

pymsgbox.alert('POS Short/Over: ' + str(pos_shortOver) + '\n' + 'Short/Over: ' + str(shortOver.value),timeout= 4000)
if float(pos_shortOver) == float(shortOver.value):
    pymsgbox.alert('Short/Over Match ✔️',timeout=3000)
else: 
    pymsgbox.alert('Short/Over Not Match ❌',timeout=3000) 













